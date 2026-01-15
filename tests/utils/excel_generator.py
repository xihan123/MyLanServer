#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 数据生成器
支持基于表头智能识别字段类型并生成相应的测试数据
"""

import random
import string
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from faker import Faker
    FAKER_AVAILABLE = True
    fake = Faker('zh_CN')
except ImportError:
    FAKER_AVAILABLE = False


class ExcelDataGenerator:
    """Excel 数据生成器"""

    # 常见中文姓氏
    SURNAMES = [
        '赵', '钱', '孙', '李', '周', '吴', '郑', '王', '冯', '陈',
        '褚', '卫', '蒋', '沈', '韩', '杨', '朱', '秦', '尤', '许',
        '何', '吕', '施', '张', '孔', '曹', '严', '华', '金', '魏'
    ]

    # 常见中文名字
    NAMES = [
        '明', '国', '华', '文', '平', '志', '伟', '芳', '军', '敏',
        '静', '强', '磊', '军', '洋', '勇', '艳', '杰', '娟', '涛',
        '超', '波', '明', '秀', '刚', '平', '辉', '鹏', '飞', '鑫'
    ]

    # 常见部门
    DEPARTMENTS = [
        '技术部', '人事部', '财务部', '市场部', '销售部',
        '运营部', '产品部', '客服部', '行政部', '法务部'
    ]

    # 常见地址
    ADDRESSES = [
        '北京市朝阳区', '上海市浦东新区', '广州市天河区', '深圳市南山区',
        '杭州市西湖区', '成都市武侯区', '武汉市江汉区', '南京市鼓楼区'
    ]

    def __init__(self, template_path: Optional[str] = None, template_content: Optional[bytes] = None):
        """
        初始化 Excel 数据生成器

        Args:
            template_path: 模板文件路径
            template_content: 模板文件内容（字节）
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError('openpyxl 库未安装，请先安装: pip install openpyxl')

        self.template_path = template_path
        self.template_content = template_content
        self.headers = []
        self.field_types = {}

        # 读取模板
        if template_path or template_content:
            self._read_template()

    def _read_template(self):
        """读取模板文件，提取表头"""
        try:
            if self.template_content:
                # 从字节内容读取
                from openpyxl import load_workbook
                wb = load_workbook(BytesIO(self.template_content))
                ws = wb.active
            else:
                # 从文件路径读取
                from openpyxl import load_workbook
                wb = load_workbook(self.template_path)
                ws = wb.active

            # 读取第一行作为表头
            for cell in ws[1]:
                if cell.value:
                    self.headers.append(str(cell.value).strip())

            # 推断每个字段的类型
            for header in self.headers:
                self.field_types[header] = self._infer_field_type(header)

        except Exception as e:
            raise Exception(f'读取模板失败: {str(e)}')

    def _infer_field_type(self, field_name: str) -> str:
        """
        根据表头名称推断字段类型

        Args:
            field_name: 字段名称

        Returns:
            字段类型
        """
        field_name_lower = field_name.lower()

        # 姓名相关
        if any(keyword in field_name for keyword in ['姓名', '人名', '用户', 'name']):
            return 'name'

        # 年龄相关
        if any(keyword in field_name for keyword in ['年龄', '岁数', 'age']):
            return 'age'

        # 电话相关
        if any(keyword in field_name for keyword in ['电话', '手机', '联系方式', 'phone', 'tel']):
            return 'phone'

        # 邮箱相关
        if any(keyword in field_name for keyword in ['邮箱', 'email', 'mail']):
            return 'email'

        # 身份证相关
        if any(keyword in field_name for keyword in ['身份证', '证件号', 'id']):
            return 'id_card'

        # 日期相关
        if any(keyword in field_name for keyword in ['日期', '时间', '生日', 'date', 'time']):
            return 'date'

        # 是否相关
        if any(keyword in field_name for keyword in ['是否', '是', '否', '启用', '禁用', 'active']):
            return 'boolean'

        # 部门相关
        if any(keyword in field_name for keyword in ['部门', '单位', 'department']):
            return 'department'

        # 地址相关
        if any(keyword in field_name for keyword in ['地址', '住址', 'address']):
            return 'address'

        # 默认为文本
        return 'text'

    def _generate_name(self) -> str:
        """生成中文姓名"""
        surname = random.choice(self.SURNAMES)
        name = random.choice(self.NAMES)
        if random.random() > 0.5:
            name += random.choice(self.NAMES)
        return surname + name

    def _generate_age(self) -> int:
        """生成年龄（18-65）"""
        return random.randint(18, 65)

    def _generate_phone(self) -> str:
        """生成手机号码"""
        # 手机号前缀
        prefixes = ['130', '131', '132', '133', '134', '135', '136', '137', '138', '139',
                    '150', '151', '152', '153', '155', '156', '157', '158', '159',
                    '180', '181', '182', '183', '184', '185', '186', '187', '188', '189']
        prefix = random.choice(prefixes)
        suffix = ''.join(random.choice(string.digits) for _ in range(8))
        return prefix + suffix

    def _generate_email(self) -> str:
        """生成邮箱地址"""
        if FAKER_AVAILABLE:
            return fake.email()
        else:
            username = ''.join(random.choice(string.ascii_lowercase) for _ in range(8))
            domains = ['qq.com', '163.com', 'gmail.com', 'outlook.com', 'hotmail.com']
            return f'{username}@{random.choice(domains)}'

    def _generate_id_card(self) -> str:
        """生成身份证号（18位）"""
        # 简化版身份证号生成（仅用于测试）
        area_code = str(random.randint(110000, 659999))
        birth_date = datetime(random.randint(1970, 2005), random.randint(1, 12), random.randint(1, 28)).strftime('%Y%m%d')
        sequence_code = ''.join(random.choice(string.digits) for _ in range(3))
        check_code = random.choice('0123456789X')
        return area_code + birth_date + sequence_code + check_code

    def _generate_date(self) -> str:
        """生成日期"""
        start_date = datetime.now() - timedelta(days=365)
        end_date = datetime.now()
        random_date = start_date + timedelta(days=random.randint(0, 365))
        return random_date.strftime('%Y-%m-%d')

    def _generate_boolean(self) -> str:
        """生成布尔值（是/否）"""
        return random.choice(['是', '否'])

    def _generate_department(self) -> str:
        """生成部门"""
        return random.choice(self.DEPARTMENTS)

    def _generate_address(self) -> str:
        """生成地址"""
        if FAKER_AVAILABLE:
            return fake.address()
        else:
            street = random.randint(1, 999)
            building = random.choice(['A', 'B', 'C', 'D'])
            room = random.randint(1, 999)
            base_address = random.choice(self.ADDRESSES)
            return f'{base_address}{street}号{building}栋{room}室'

    def _generate_text(self, length: int = 10) -> str:
        """生成随机文本"""
        if FAKER_AVAILABLE:
            return fake.sentence()[:length]
        else:
            return ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(length))

    def generate_row(self) -> Dict[str, Any]:
        """
        生成一行数据

        Returns:
            字段名到值的映射
        """
        row_data = {}

        for header in self.headers:
            field_type = self.field_types.get(header, 'text')

            # 根据字段类型生成数据
            if field_type == 'name':
                value = self._generate_name()
            elif field_type == 'age':
                value = self._generate_age()
            elif field_type == 'phone':
                value = self._generate_phone()
            elif field_type == 'email':
                value = self._generate_email()
            elif field_type == 'id_card':
                value = self._generate_id_card()
            elif field_type == 'date':
                value = self._generate_date()
            elif field_type == 'boolean':
                value = self._generate_boolean()
            elif field_type == 'department':
                value = self._generate_department()
            elif field_type == 'address':
                value = self._generate_address()
            else:  # text
                value = self._generate_text()

            row_data[header] = value

        return row_data

    def generate_rows(self, count: int) -> List[Dict[str, Any]]:
        """
        生成多行数据

        Args:
            count: 行数

        Returns:
            数据列表
        """
        return [self.generate_row() for _ in range(count)]

    def generate_excel_file(self, data: List[Dict[str, Any]], include_header: bool = True) -> bytes:
        """
        生成 Excel 文件

        Args:
            data: 数据列表
            include_header: 是否包含表头

        Returns:
            Excel 文件的字节内容
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError('openpyxl 库未安装，请先安装: pip install openpyxl')

        wb = Workbook()
        ws = wb.active
        ws.title = "数据"

        # 写入表头
        if include_header and self.headers:
            for col_idx, header in enumerate(self.headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

        # 写入数据
        start_row = 2 if include_header else 1
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, header in enumerate(self.headers, start=1):
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 保存到字节流
        buffer = BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        return content

    def get_headers(self) -> List[str]:
        """
        获取表头列表

        Returns:
            表头列表
        """
        return self.headers

    def get_field_types(self) -> Dict[str, str]:
        """
        获取字段类型映射

        Returns:
            字段名到类型的映射
        """
        return self.field_types

    @staticmethod
    def create_from_template(template_path: str) -> 'ExcelDataGenerator':
        """
        从模板文件创建生成器

        Args:
            template_path: 模板文件路径

        Returns:
            Excel 数据生成器实例
        """
        return ExcelDataGenerator(template_path=template_path)

    @staticmethod
    def create_from_template_content(template_content: bytes) -> 'ExcelDataGenerator':
        """
        从模板内容创建生成器

        Args:
            template_content: 模板文件内容（字节）

        Returns:
            Excel 数据生成器实例
        """
        return ExcelDataGenerator(template_content=template_content)